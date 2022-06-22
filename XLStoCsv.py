__title__ = "XLStoCsv"
__version__ = "0.0.1"
__author__ = "Alexandre Heneffe"
__license__ = "MIT"
__copyright__ = "Copyright 2022 Alexandre Heneffe"

"""
Simple script to convert a xlsx (Excel) file to a csv file
"""

import pandas as pd
import xlrd
from openpyxl import load_workbook


xls_file_path = "sensors/FluksoTechnical.xlsx"
csv_file_path = "sensors/FluksoTechnical_Flukso.csv"


def convertXLStoCSV():
    read_file = pd.read_excel(xls_file_path, sheet_name="Flukso")
    print(read_file.head(10))
    read_file.to_csv(csv_file_path, index=None, header=True)


def testXLSreader():
    # load excel file
    workbook = load_workbook(filename=xls_file_path)

    # open workbook
    sheet = workbook["Sensors"]

    print("nb of rows : ", sheet.max_row)
    for i in range(1, sheet.max_row+1):
        print(sheet.cell(row=i, column=1).value)

    # modify the desired cell
    # print(sheet.cell(row=1, column=1).value)
    # sheet["A1"] = "Full Name"

    # save the file
    # workbook.save(filename="csv/output.xlsx")


def main():
    # convertXLStoCSV()

    testXLSreader()


if __name__ == "__main__":
    main()


